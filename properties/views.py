from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from datetime import timedelta, date
from .models import Unit, Customer, Cicilan, Perumahan, CompanyInfo
from .forms import CustomerForm, UnitForm, CustomerRegistrationForm, ConfirmasiBayarForm, CompanyInfoForm
from django.db.models import Sum
from django.http import HttpResponse
from django.contrib import messages
import openpyxl
import calendar
from decimal import Decimal

def add_months(sourcedate, months):
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, calendar.monthrange(year, month)[1])
    return sourcedate.replace(year=year, month=month, day=day)

def auto_update_sp(cicilan, today):
    """Auto-update status SP berdasarkan keterlambatan"""
    if cicilan.status_bayar == 'Belum Lunas' and cicilan.tanggal_batas_bayar:
        hari_terlambat = (today - cicilan.tanggal_batas_bayar).days
        if hari_terlambat >= 7:
            if cicilan.status_sp != 'SP1':
                cicilan.status_sp = 'SP1'
                cicilan.save(update_fields=['status_sp'])
    return cicilan

def dashboard(request):
    filter_perumahan_id = request.GET.get('perumahan_id', '')
    
    units_qs = Unit.objects.all()
    if filter_perumahan_id:
        units_qs = units_qs.filter(perumahan_id=filter_perumahan_id)
        
    total_unit = units_qs.count()
    unit_tersedia = units_qs.filter(status='Tersedia').count()
    unit_terjual_atau_booking = total_unit - unit_tersedia
    
    today = timezone.now().date()
    batas_waktu = today + timedelta(days=7)
    
    cicilan_jatuh_tempo_qs = Cicilan.objects.filter(
        status_bayar='Belum Lunas',
        tanggal_jatuh_tempo__lte=batas_waktu
    ).select_related('customer', 'unit__perumahan')
    if filter_perumahan_id:
        cicilan_jatuh_tempo_qs = cicilan_jatuh_tempo_qs.filter(unit__perumahan_id=filter_perumahan_id)
    
    # Auto-update status SP dan hitung denda
    cicilan_jatuh_tempo_list = []
    for c in cicilan_jatuh_tempo_qs.order_by('tanggal_jatuh_tempo'):
        c = auto_update_sp(c, today)
        cicilan_jatuh_tempo_list.append(c)
        
    import json
    months_revenue = [0] * 12
    months_sales_qty = [0] * 12
    months_sales_value = [0] * 12
    
    lunas_this_year = Cicilan.objects.filter(status_bayar='Lunas', tahun=today.year)
    if filter_perumahan_id:
        lunas_this_year = lunas_this_year.filter(unit__perumahan_id=filter_perumahan_id)
        
    for c in lunas_this_year:
        if 1 <= c.bulan <= 12:
            months_revenue[c.bulan - 1] += int(c.jumlah_cicilan)
            
    sold_units_qs = Unit.objects.exclude(status='Tersedia').prefetch_related('cicilan')
    if filter_perumahan_id:
        sold_units_qs = sold_units_qs.filter(perumahan_id=filter_perumahan_id)
        
    for u in sold_units_qs:
        first_installment = u.cicilan.order_by('tanggal_jatuh_tempo', 'id').first()
        if first_installment:
            if first_installment.tanggal_jatuh_tempo.year == today.year:
                sale_month = first_installment.tanggal_jatuh_tempo.month
                if 1 <= sale_month <= 12:
                    months_sales_qty[sale_month - 1] += 1
                    months_sales_value[sale_month - 1] += int(u.harga_total)
            
    perumahans = Perumahan.objects.all().order_by('nama_perumahan')
    
    context = {
        'total_unit': total_unit,
        'unit_tersedia': unit_tersedia,
        'unit_terjual': unit_terjual_atau_booking,
        'cicilan_jatuh_tempo': cicilan_jatuh_tempo_list,
        'today': today,
        'revenue_data': json.dumps(months_revenue),
        'sales_qty_data': json.dumps(months_sales_qty),
        'sales_value_data': json.dumps(months_sales_value),
        'current_year': today.year,
        'perumahans': perumahans,
        'filter_perumahan_id': int(filter_perumahan_id) if filter_perumahan_id else '',
    }
    return render(request, 'properties/dashboard.html', context)

def export_cicilan_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Laporan_Cicilan_Jatuh_Tempo.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Cicilan Jatuh Tempo'

    columns = ['No', 'Blok/Unit', 'Nama Customer', 'No. HP', 'Jumlah Cicilan (Rp)', 'Jatuh Tempo', 'Batas Bayar', 'Periode', 'Keterangan', 'Denda (Rp)', 'Status SP', 'Status']
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=1, column=col_num, value=column_title)

    today = timezone.now().date()
    batas_waktu = today + timedelta(days=7)
    
    cicilan_jatuh_tempo = Cicilan.objects.filter(
        status_bayar='Belum Lunas',
        tanggal_jatuh_tempo__lte=batas_waktu
    ).order_by('tanggal_jatuh_tempo').select_related('customer', 'unit')

    for idx, item in enumerate(cicilan_jatuh_tempo, 1):
        status = "Terlewat" if item.tanggal_jatuh_tempo < today else "Hampir Tiba"
        no_hp = item.customer.no_telepon if item.customer.no_telepon else "-"
        batas = item.tanggal_batas_bayar.strftime('%d-%m-%Y') if item.tanggal_batas_bayar else "-"
        denda_val = float(item.denda_terhitung) if item.denda_terhitung else 0
        
        row = [idx, item.unit.kode_blok, item.customer.nama_lengkap, no_hp,
               float(item.jumlah_cicilan), item.tanggal_jatuh_tempo.strftime('%d-%m-%Y'),
               batas, f"{item.bulan}/{item.tahun}", item.keterangan_cicilan,
               denda_val, item.status_sp, status]
        
        for col_num, cell_value in enumerate(row, 1):
            worksheet.cell(row=idx + 1, column=col_num, value=cell_value)

    workbook.save(response)
    return response

# --- RIWAYAT SEMUA CICILAN (BISA UPDATE KAPAN SAJA) ---

def riwayat_cicilan(request):
    search_query = request.GET.get('search', '')
    filter_perumahan_id = request.GET.get('perumahan_id', '')
    filter_status = request.GET.get('status', '')
    
    cicilan_qs = Cicilan.objects.select_related(
        'customer', 'unit__perumahan'
    ).order_by('customer__nama_lengkap', 'tanggal_jatuh_tempo')
    
    if search_query:
        from django.db.models import Q
        cicilan_qs = cicilan_qs.filter(
            Q(customer__nama_lengkap__icontains=search_query) |
            Q(customer__no_telepon__icontains=search_query)
        )
    
    if filter_perumahan_id:
        cicilan_qs = cicilan_qs.filter(unit__perumahan_id=filter_perumahan_id)
    
    if filter_status:
        cicilan_qs = cicilan_qs.filter(status_bayar=filter_status)
    
    today = timezone.now().date()
    
    # Auto-update SP status
    for c in cicilan_qs:
        auto_update_sp(c, today)
    
    perumahans = Perumahan.objects.all().order_by('nama_perumahan')
    
    context = {
        'cicilans': cicilan_qs,
        'title': 'Riwayat Seluruh Cicilan',
        'perumahans': perumahans,
        'search_query': search_query,
        'filter_perumahan_id': int(filter_perumahan_id) if filter_perumahan_id else '',
        'filter_status': filter_status,
        'today': today,
    }
    return render(request, 'properties/riwayat_cicilan.html', context)


# --- KONFIRMASI & AKSI PEMBAYARAN CICILAN ---

def konfirmasi_bayar(request, pk):
    cicilan = get_object_or_404(Cicilan, pk=pk)
    source = request.GET.get('from', 'riwayat')
    customer = cicilan.customer
    saldo_kredit = customer.saldo_kredit or Decimal('0')
    tagihan_pokok = cicilan.jumlah_cicilan
    
    # Hitung tagihan setelah potong saldo kredit
    if saldo_kredit >= tagihan_pokok:
        tagihan_setelah_kredit = Decimal('0')
    else:
        tagihan_setelah_kredit = tagihan_pokok - saldo_kredit
    
    if request.method == 'POST':
        form = ConfirmasiBayarForm(request.POST, instance=cicilan)
        if form.is_valid():
            jumlah_dibayar = form.cleaned_data['jumlah_dibayar']
            metode_bayar = form.cleaned_data['metode_bayar']
            rekening = form.cleaned_data.get('rekening', '-')
            
            old_ket = cicilan.keterangan_cicilan
            
            # Total pembayaran efektif = saldo kredit yang dipakai + uang tunai dibayar
            kredit_dipakai = min(saldo_kredit, tagihan_pokok)
            total_bayar_efektif = kredit_dipakai + jumlah_dibayar
            
            cicilan.jumlah_dibayar = jumlah_dibayar
            cicilan.metode_bayar = metode_bayar
            cicilan.rekening = rekening if metode_bayar == 'Transfer' else '-'
            cicilan.status_bayar = 'Lunas'
            if not cicilan.keterangan_cicilan.endswith("(Lunas)"):
                cicilan.keterangan_cicilan = f"{cicilan.keterangan_cicilan} (Lunas)"
            cicilan.save()
            
            # Potong saldo kredit yang terpakai
            if kredit_dipakai > 0:
                customer.saldo_kredit = saldo_kredit - kredit_dipakai
                customer.save(update_fields=['saldo_kredit'])
            
            # Handle kelebihan bayar baru → tambah saldo kredit
            selisih = total_bayar_efektif - tagihan_pokok
            if selisih > 0:
                customer.saldo_kredit = (customer.saldo_kredit or Decimal('0')) + selisih
                customer.save(update_fields=['saldo_kredit'])
                messages.success(request, 
                    f"Pembayaran {old_ket} a.n {customer.nama_lengkap} lunas via {metode_bayar}. "
                    f"Saldo kredit dipakai: Rp {kredit_dipakai:,.0f}. "
                    f"Kelebihan bayar Rp {selisih:,.0f} masuk ke Saldo Kredit."
                )
            else:
                msg = f"Pembayaran {old_ket} a.n {customer.nama_lengkap} berhasil divalidasi lunas via {metode_bayar}."
                if kredit_dipakai > 0:
                    msg += f" Saldo kredit dipakai: Rp {kredit_dipakai:,.0f}."
                messages.success(request, msg)
            
            if source == 'dashboard':
                return redirect('dashboard')
            return redirect('riwayat_cicilan')
    else:
        # Pre-fill jumlah_dibayar = tagihan - saldo kredit
        form = ConfirmasiBayarForm(instance=cicilan, initial={'jumlah_dibayar': tagihan_setelah_kredit})
    
    return render(request, 'properties/konfirmasi_bayar.html', {
        'form': form,
        'cicilan': cicilan,
        'source': source,
        'saldo_kredit': saldo_kredit,
        'tagihan_setelah_kredit': tagihan_setelah_kredit,
    })

# --- REKAP STATUS SEMUA KONSUMEN ---

def status_konsumen(request):
    search_query = request.GET.get('search', '')
    filter_perumahan_id = request.GET.get('perumahan_id', '')

    customers_qs = Customer.objects.prefetch_related('cicilan__unit__perumahan').order_by('nama_lengkap')
    
    if search_query:
        from django.db.models import Q
        customers_qs = customers_qs.filter(
            Q(nama_lengkap__icontains=search_query) | 
            Q(no_telepon__icontains=search_query)
        )
        
    customers = customers_qs.distinct()
    cicilans_to_display = []
    
    for customer in customers:
        cicilan_qs = customer.cicilan.all()
        
        if filter_perumahan_id:
            cicilan_qs = cicilan_qs.filter(unit__perumahan_id=filter_perumahan_id)
            
        if not cicilan_qs.exists():
            continue
            
        next_unpaid = cicilan_qs.filter(status_bayar='Belum Lunas').order_by('tanggal_jatuh_tempo').first()
        
        if next_unpaid:
            terbayar_agg = cicilan_qs.filter(
                unit=next_unpaid.unit,
                status_bayar='Lunas'
            ).aggregate(total=Sum('jumlah_cicilan'))
            
            next_unpaid.total_terbayar = terbayar_agg['total'] or 0
            
            semua_cicilan_agg = cicilan_qs.filter(
                unit=next_unpaid.unit
            ).aggregate(total=Sum('jumlah_cicilan'))
            next_unpaid.harga_rumah = semua_cicilan_agg['total'] or 0
            next_unpaid.sisa_hutang = next_unpaid.harga_rumah - next_unpaid.total_terbayar
            
            # Total harga full = harga properti + bunga (bunga dihitung dari sisa setelah DP/UTJ)
            harga_unit = next_unpaid.unit.harga_total or Decimal('0')
            dp = customer.dp or Decimal('0')
            utj = customer.utj or Decimal('0')
            sisa_setelah_dp_utj = harga_unit - dp - utj
            bunga_persen = customer.bunga_per_tahun or Decimal('0')
            lama_bulan = cicilan_qs.filter(unit=next_unpaid.unit).count()
            lama_tahun = Decimal(str(lama_bulan)) / Decimal('12')
            total_bunga = sisa_setelah_dp_utj * bunga_persen / Decimal('100') * lama_tahun
            next_unpaid.total_harga_full = harga_unit + total_bunga
            next_unpaid.dp_customer = dp
            next_unpaid.utj_customer = utj
            next_unpaid.total_bunga = total_bunga
            
            cicilans_to_display.append(next_unpaid)
        else:
            first_paid = cicilan_qs.filter(status_bayar='Lunas').order_by('tanggal_jatuh_tempo').last()
            if first_paid:
                semua_cicilan_agg = cicilan_qs.filter(
                    unit=first_paid.unit
                ).aggregate(total=Sum('jumlah_cicilan'))
                total_semua = semua_cicilan_agg['total'] or 0
                
                first_paid.keterangan_cicilan = "LUNAS SEMUA"
                first_paid.total_terbayar = total_semua
                first_paid.harga_rumah = total_semua
                first_paid.sisa_hutang = 0
                first_paid.status_bayar = 'Lunas'
                
                harga_unit = first_paid.unit.harga_total or Decimal('0')
                dp = customer.dp or Decimal('0')
                utj = customer.utj or Decimal('0')
                sisa_setelah_dp_utj = harga_unit - dp - utj
                bunga_persen = customer.bunga_per_tahun or Decimal('0')
                lama_bulan = cicilan_qs.filter(unit=first_paid.unit).count()
                lama_tahun = Decimal(str(lama_bulan)) / Decimal('12')
                total_bunga = sisa_setelah_dp_utj * bunga_persen / Decimal('100') * lama_tahun
                first_paid.total_harga_full = harga_unit + total_bunga
                first_paid.dp_customer = dp
                first_paid.utj_customer = utj
                first_paid.total_bunga = total_bunga
                
                cicilans_to_display.append(first_paid)
                
    perumahans = Perumahan.objects.all().order_by('nama_perumahan')

    context_data = {
        'cicilans': cicilans_to_display, 
        'title': 'Data Status Cicilan Konsumen',
        'perumahans': perumahans,
        'search_query': search_query,
        'filter_perumahan_id': int(filter_perumahan_id) if filter_perumahan_id else ''
    }
    return render(request, 'properties/status_konsumen.html', context_data)

def export_konsumen_excel(request):
    search_query = request.GET.get('search', '')
    filter_perumahan_id = request.GET.get('perumahan_id', '')

    customers_qs = Customer.objects.prefetch_related('cicilan__unit__perumahan').order_by('nama_lengkap')
    
    if search_query:
        from django.db.models import Q
        customers_qs = customers_qs.filter(
            Q(nama_lengkap__icontains=search_query) | 
            Q(no_telepon__icontains=search_query)
        )
        
    customers = customers_qs.distinct()
    cicilans_to_display = []
    
    for customer in customers:
        cicilan_qs = customer.cicilan.all()
        if filter_perumahan_id:
            cicilan_qs = cicilan_qs.filter(unit__perumahan_id=filter_perumahan_id)
            
        if not cicilan_qs.exists():
            continue
            
        next_unpaid = cicilan_qs.filter(status_bayar='Belum Lunas').order_by('tanggal_jatuh_tempo').first()
        if next_unpaid:
            terbayar_agg = cicilan_qs.filter(
                unit=next_unpaid.unit,
                status_bayar='Lunas'
            ).aggregate(total=Sum('jumlah_cicilan'))
            next_unpaid.total_terbayar = terbayar_agg['total'] or 0
            
            semua_cicilan_agg = cicilan_qs.filter(
                unit=next_unpaid.unit
            ).aggregate(total=Sum('jumlah_cicilan'))
            next_unpaid.harga_rumah = semua_cicilan_agg['total'] or 0
            
            next_unpaid.sisa_hutang = next_unpaid.harga_rumah - next_unpaid.total_terbayar
            cicilans_to_display.append(next_unpaid)
        else:
            first_paid = cicilan_qs.filter(status_bayar='Lunas').order_by('tanggal_jatuh_tempo').last()
            if first_paid:
                semua_cicilan_agg = cicilan_qs.filter(
                    unit=first_paid.unit
                ).aggregate(total=Sum('jumlah_cicilan'))
                total_semua = semua_cicilan_agg['total'] or 0
                
                first_paid.keterangan_cicilan = "LUNAS SEMUA"
                first_paid.total_terbayar = total_semua
                first_paid.harga_rumah = total_semua
                first_paid.sisa_hutang = 0
                cicilans_to_display.append(first_paid)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Rekap_Status_Konsumen.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data Konsumen'

    columns = ['No', 'NIK', 'Nama Customer', 'No. Telp', 'Metode Beli', 'Perumahan', 'Blok/Unit', 'Cicilan Ke', 'Jlh Tagihan (Rp)', 'Total Terbayar (Rp)', 'Sisa Hutang (Rp)', 'Saldo Kredit (Rp)']
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=1, column=col_num, value=column_title)

    row_num = 1
    for item in cicilans_to_display:
        row_num += 1
        perumahan = item.unit.perumahan.nama_perumahan if item.unit.perumahan else "-"
        blok_unit = f"{item.unit.kode_blok} - {item.unit.tipe_rumah}"
        telp = item.customer.no_telepon if item.customer.no_telepon else "-"
        nik = item.customer.nik if item.customer.nik else "-"
        metode = item.customer.metode_pembelian if item.customer.metode_pembelian else "-"
        saldo = float(item.customer.saldo_kredit) if item.customer.saldo_kredit else 0
        
        row = [row_num - 1, nik, item.customer.nama_lengkap, telp, metode, perumahan, blok_unit,
               item.keterangan_cicilan, float(item.jumlah_cicilan),
               float(item.total_terbayar), float(item.sisa_hutang), saldo]
        
        for col_num, cell_value in enumerate(row, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(response)
    return response

# --- CRUD CUSTOMER ---

def customer_create(request):
    if request.method == 'POST':
        form = CustomerRegistrationForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            
            unit = form.cleaned_data['unit']
            metode_pembelian = form.cleaned_data['metode_pembelian']
            utj = form.cleaned_data.get('utj') or Decimal('0')
            dp = form.cleaned_data.get('dp') or Decimal('0')
            bunga_per_tahun = form.cleaned_data.get('bunga_per_tahun') or Decimal('0')
            lama_cicilan = form.cleaned_data['lama_cicilan']
            tanggal_jatuh_tempo = form.cleaned_data['tanggal_jatuh_tempo']
            offset_hari = form.cleaned_data.get('tanggal_batas_bayar_offset') or 7
            denda_persen = form.cleaned_data.get('denda_persen') or Decimal('0')

            # Simpan info pembelian ke customer
            customer.metode_pembelian = metode_pembelian
            customer.utj = utj
            customer.dp = dp
            customer.bunga_per_tahun = bunga_per_tahun
            customer.save()

            # --- Kalkulasi Cicilan ---
            harga_unit = unit.harga_total
            pokok = harga_unit - utj - dp

            if metode_pembelian == 'KPR':
                bunga_per_bulan = pokok * (bunga_per_tahun / Decimal('100')) / Decimal('12')
                jumlah_cicilan = (pokok / lama_cicilan) + bunga_per_bulan
            else:
                # Cash Bertahap & Cash Keras: logika sama, tanpa bunga
                jumlah_cicilan = pokok / lama_cicilan

            # Update status unit
            if unit.status == 'Tersedia':
                unit.status = 'Booking'
                unit.save()
                
            # Bulk create cicilan
            cicilan_bulk = []
            for i in range(lama_cicilan):
                tgl_jt = add_months(tanggal_jatuh_tempo, i)
                tgl_batas = tgl_jt + timedelta(days=offset_hari)
                cicilan_bulk.append(Cicilan(
                    customer=customer,
                    unit=unit,
                    jumlah_cicilan=jumlah_cicilan,
                    tanggal_jatuh_tempo=tgl_jt,
                    tanggal_batas_bayar=tgl_batas,
                    bulan=tgl_jt.month,
                    tahun=tgl_jt.year,
                    keterangan_cicilan=f"C{i+1}",
                    denda_persen=denda_persen,
                    rekening="-",
                    status_bayar="Belum Lunas"
                ))
            
            if cicilan_bulk:
                Cicilan.objects.bulk_create(cicilan_bulk)

            messages.success(request, 
                f"Pelanggan {customer.nama_lengkap} terdaftar ({metode_pembelian}). "
                f"Cicilan C1–C{lama_cicilan} @ Rp {jumlah_cicilan:,.0f}/bln dibuat otomatis."
            )
            return redirect('status_konsumen')
    else:
        form = CustomerRegistrationForm()
        
    import json
    units = Unit.objects.filter(status='Tersedia').select_related('perumahan')
    perumahan_units = {}
    
    for u in units:
        p_id = u.perumahan_id if u.perumahan_id else "null"
        if p_id not in perumahan_units:
            perumahan_units[p_id] = []
            
        perumahan_units[p_id].append({
            'id': u.id,
            'text': f"{u.kode_blok} - {u.tipe_rumah}",
            'price': float(u.harga_total)
        })
        
    return render(request, 'properties/customer_form.html', {
        'form': form, 
        'title': 'Register Pelanggan Baru & Cicilan',
        'perumahan_units_json': json.dumps(perumahan_units)
    })

def customer_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    cicilans = customer.cicilan.select_related('unit').all()
    units_owned = list(set([c.unit for c in cicilans]))
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f"Data Customer {customer.nama_lengkap} berhasil diperbarui.")
            return redirect('status_konsumen')
    else:
        initial_data = {}
        if not customer.alamat_sekarang and units_owned:
            bloks = ", ".join([u.kode_blok for u in units_owned])
            initial_data['alamat_sekarang'] = f"Blok {bloks}"
            
        form = CustomerForm(instance=customer, initial=initial_data)
        
    return render(request, 'properties/customer_form.html', {'form': form, 'title': 'Edit Data Customer', 'units_owned': units_owned})

def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        nama = customer.nama_lengkap
        customer.delete()
        messages.success(request, f"Data Customer {nama} berhasil dihapus beserta seluruh cicilannya.")
        return redirect('status_konsumen')
    return render(request, 'properties/customer_confirm_delete.html', {'customer': customer})

# --- CRUD UNIT / PROPERTI ---

def unit_list(request):
    units = Unit.objects.select_related('perumahan').all().order_by('perumahan__nama_perumahan', 'blok', 'nomor_rumah')
    return render(request, 'properties/unit_list.html', {'units': units})

def export_properti_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Rekap_Data_Properti.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data Properti'

    columns = ['No', 'Perumahan', 'Blok', 'No. Rumah', 'Tipe', 'Luas Tanah (m²)', 'Luas Bangunan (m²)', 'Harga Total (Rp)', 'Status']
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=1, column=col_num, value=column_title)

    units = Unit.objects.select_related('perumahan').all().order_by('perumahan__nama_perumahan', 'blok', 'nomor_rumah')
    
    for idx, item in enumerate(units, 1):
        perumahan = item.perumahan.nama_perumahan if item.perumahan else "-"
        lt = float(item.luas_tanah) if item.luas_tanah else "-"
        lb = float(item.luas_bangunan) if item.luas_bangunan else "-"
        row = [idx, perumahan, item.blok, item.nomor_rumah, item.tipe_rumah, lt, lb, float(item.harga_total), item.status]
        for col_num, cell_value in enumerate(row, 1):
            worksheet.cell(row=idx + 1, column=col_num, value=cell_value)

    workbook.save(response)
    return response

def unit_create(request):
    if request.method == 'POST':
        form = UnitForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Data Properti/Unit berhasil ditambahkan.")
            return redirect('unit_list')
    else:
        form = UnitForm()
    return render(request, 'properties/unit_form.html', {'form': form, 'title': 'Tambah Properti/Unit'})

def unit_update(request, pk):
    unit = get_object_or_404(Unit, pk=pk)
    if request.method == 'POST':
        form = UnitForm(request.POST, instance=unit)
        if form.is_valid():
            form.save()
            messages.success(request, f"Data Properti/Unit {unit.kode_blok} berhasil diperbarui.")
            return redirect('unit_list')
    else:
        form = UnitForm(instance=unit)
    return render(request, 'properties/unit_form.html', {'form': form, 'title': 'Edit Properti/Unit'})

def unit_delete(request, pk):
    unit = get_object_or_404(Unit, pk=pk)
    if request.method == 'POST':
        kode = unit.kode_blok
        unit.delete()
        messages.success(request, f"Data Properti/Unit {kode} berhasil dihapus.")
        return redirect('unit_list')
    return render(request, 'properties/unit_confirm_delete.html', {'unit': unit})

# --- COMPANY SETTINGS ---

def company_settings(request):
    company = CompanyInfo.load()
    perumahans = Perumahan.objects.all().order_by('nama_perumahan')
    
    if request.method == 'POST':
        form_type = request.POST.get('form_type', 'company')
        
        if form_type == 'perumahan':
            # Update desa/kelurahan & kecamatan untuk semua perumahan
            for p in perumahans:
                desa = request.POST.get(f'desa_{p.pk}', '').strip()
                kec = request.POST.get(f'kecamatan_{p.pk}', '').strip()
                if desa != (p.desa_kelurahan or '') or kec != (p.kecamatan or ''):
                    p.desa_kelurahan = desa
                    p.kecamatan = kec
                    p.save()
            messages.success(request, "Data lokasi perumahan berhasil disimpan.")
            return redirect('company_settings')
        else:
            form = CompanyInfoForm(request.POST, instance=company)
            if form.is_valid():
                form.save()
                messages.success(request, "Pengaturan perusahaan berhasil disimpan.")
                return redirect('company_settings')
    else:
        form = CompanyInfoForm(instance=company)
    
    return render(request, 'properties/company_settings.html', {
        'form': form,
        'perumahans': perumahans,
    })

# --- TERBILANG (Angka ke Kata Indonesia) ---

def terbilang(n):
    """Convert number to Indonesian words"""
    n = int(n)
    if n == 0:
        return "Nol"
    
    satuan = ['', 'Satu', 'Dua', 'Tiga', 'Empat', 'Lima', 'Enam', 'Tujuh', 'Delapan', 'Sembilan', 'Sepuluh', 'Sebelas']
    
    if n < 0:
        return "Minus " + terbilang(-n)
    elif n < 12:
        return satuan[n]
    elif n < 20:
        return terbilang(n - 10) + " Belas"
    elif n < 100:
        return terbilang(n // 10) + " Puluh" + (" " + terbilang(n % 10) if n % 10 else "")
    elif n < 200:
        return "Seratus" + (" " + terbilang(n - 100) if n - 100 else "")
    elif n < 1000:
        return terbilang(n // 100) + " Ratus" + (" " + terbilang(n % 100) if n % 100 else "")
    elif n < 2000:
        return "Seribu" + (" " + terbilang(n - 1000) if n - 1000 else "")
    elif n < 1000000:
        return terbilang(n // 1000) + " Ribu" + (" " + terbilang(n % 1000) if n % 1000 else "")
    elif n < 1000000000:
        return terbilang(n // 1000000) + " Juta" + (" " + terbilang(n % 1000000) if n % 1000000 else "")
    elif n < 1000000000000:
        return terbilang(n // 1000000000) + " Miliar" + (" " + terbilang(n % 1000000000) if n % 1000000000 else "")
    else:
        return terbilang(n // 1000000000000) + " Triliun" + (" " + terbilang(n % 1000000000000) if n % 1000000000000 else "")

# --- GENERATE SURAT PESANAN KAVLING (PDF) ---

def generate_surat_pesanan(request, pk):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import io
    
    customer = get_object_or_404(Customer, pk=pk)
    cicilans = Cicilan.objects.filter(customer=customer).select_related('unit__perumahan').order_by('tanggal_jatuh_tempo')
    
    if not cicilans.exists():
        messages.error(request, "Customer ini belum memiliki data cicilan.")
        return redirect('status_konsumen')
    
    unit = cicilans.first().unit
    perumahan = unit.perumahan
    company = CompanyInfo.load()
    
    # Calculations
    harga_unit = unit.harga_total or Decimal('0')
    utj = customer.utj or Decimal('0')
    dp = customer.dp or Decimal('0')
    total_dp = utj + dp
    pokok = harga_unit - utj - dp
    bunga_persen = customer.bunga_per_tahun or Decimal('0')
    lama_bulan = cicilans.count()
    lama_tahun = Decimal(str(lama_bulan)) / Decimal('12')
    
    if customer.metode_pembelian == 'KPR':
        total_bunga = pokok * bunga_persen / Decimal('100') * lama_tahun
    else:
        total_bunga = Decimal('0')
    
    harga_jual_total = harga_unit + total_bunga
    total_sisa_cicilan = pokok + total_bunga
    
    today = timezone.now().date()
    nomor_surat = f"{customer.pk:03d}/GR/SPK/{today.year}"
    
    # --- Build compact 1-page PDF ---
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                            leftMargin=1.5*cm, rightMargin=1.5*cm, 
                            topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    page_w = 18*cm  # usable width
    
    # Compact styles
    s_title = ParagraphStyle('T', parent=styles['Heading1'], fontSize=12, alignment=TA_CENTER, spaceAfter=0, spaceBefore=0)
    s_sub = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, spaceAfter=1, spaceBefore=0)
    s_n = ParagraphStyle('N', parent=styles['Normal'], fontSize=8, leading=11, spaceBefore=0, spaceAfter=0)
    s_sec = ParagraphStyle('Sec', parent=styles['Normal'], fontSize=9, leading=12, spaceBefore=0, spaceAfter=0)
    s_sec.fontName = 'Helvetica-Bold'
    s_sm = ParagraphStyle('Sm', parent=styles['Normal'], fontSize=7, leading=9)
    s_center = ParagraphStyle('Ct', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, leading=11)
    
    fmt = lambda v: f"Rp {int(v):,.0f}.-".replace(",", ".") if v else "-"
    bulan_indo = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des']
    
    kv_w = 3*cm
    val_w = page_w - kv_w - 0.4*cm
    
    def kv_table(data):
        t = Table(data, colWidths=[kv_w, 0.4*cm, val_w])
        t.setStyle(TableStyle([
            ('FONT', (0,0), (-1,-1), 'Helvetica', 8),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 0.5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0.5),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ]))
        return t
    
    el = []
    
    # === HEADER ===
    el.append(Paragraph(f"<b>{company.nama_perusahaan}</b>", s_title))
    el.append(Paragraph("<b>SURAT PESANAN KAVLING</b>", s_sub))
    el.append(Paragraph(f"NO : {nomor_surat}", s_sub))
    el.append(Spacer(1, 4))
    
    # === A. DATA PEMBELI ===
    el.append(Paragraph("<b>A. DATA PEMBELI</b>", s_sec))
    el.append(Spacer(1, 2))
    el.append(kv_table([
        ['Nama', ':', customer.nama_lengkap or '-'],
        ['No. KTP', ':', customer.nik or '-'],
        ['Alamat KTP', ':', customer.alamat_ktp or '-'],
        ['Alamat Sekarang', ':', customer.alamat_sekarang or '-'],
        ['Telepon', ':', customer.no_telepon or '-'],
    ]))
    
    if customer.nama_kontak_darurat:
        el.append(Spacer(1, 2))
        el.append(Paragraph("<b>Kontak Darurat:</b>", s_n))
        el.append(kv_table([
            ['Nama', ':', f"{customer.nama_kontak_darurat or '-'} ({customer.hubungan_kontak_darurat or '-'})"],
            ['Telepon', ':', customer.no_kontak_darurat or '-'],
        ]))
    
    el.append(Spacer(1, 4))
    
    # === B. DATA UMUM ===
    el.append(Paragraph("<b>B. DATA UMUM</b>", s_sec))
    el.append(Paragraph("<b>1. TANAH DAN BANGUNAN</b>", s_n))
    el.append(Spacer(1, 1))
    
    luas_tanah = int(unit.luas_tanah) if unit.luas_tanah else '-'
    
    el.append(kv_table([
        ['Lokasi', ':', perumahan.nama_perumahan if perumahan else '-'],
        ['Blok/Nomor', ':', f"{unit.blok.upper()}/{unit.nomor_rumah}"],
        ['Tipe', ':', f"{luas_tanah}"],
        ['Desa/Kelurahan', ':', perumahan.desa_kelurahan or '-'],
        ['Kecamatan', ':', perumahan.kecamatan or '-'],
        ['Luas Tanah', ':', f"{luas_tanah} m\u00b2" if unit.luas_tanah else '-'],
        ['Jenis Bangunan', ':', unit.tipe_rumah or '-'],
    ]))
    el.append(Spacer(1, 3))
    
    # === 2. HARGA JUAL ===
    el.append(Paragraph("<b>2. HARGA JUAL</b>", s_n))
    el.append(Spacer(1, 1))
    
    harga_str = f"Rp. {int(harga_jual_total):,.0f}.-".replace(",", ".")
    terbilang_str = f"# {terbilang(int(harga_jual_total))} Rupiah #"
    
    t_h = Table([
        ['Harga Jual', ':', harga_str],
        ['Terbilang', ':', terbilang_str],
    ], colWidths=[kv_w, 0.4*cm, val_w])
    t_h.setStyle(TableStyle([
        ('FONT', (0,0), (-1,-1), 'Helvetica', 8),
        ('FONT', (2,0), (2,0), 'Helvetica-Bold', 8),
        ('FONT', (2,1), (2,1), 'Helvetica-BoldOblique', 7.5),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 0.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0.5),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
    ]))
    el.append(t_h)
    el.append(Spacer(1, 3))
    
    # === 3. JADWAL PEMBAYARAN (2 tabel sejajar) ===
    el.append(Paragraph("<b>3. CARA DAN JADWAL PEMBAYARAN</b>", s_n))
    el.append(Spacer(1, 2))
    
    cicilan_list = list(cicilans)
    first_cicilan = cicilans.first()
    tgl_utj = first_cicilan.tanggal_jatuh_tempo if first_cicilan else today
    
    # LEFT TABLE: Jadwal Angsuran
    schedule_data = [
        [Paragraph('<b>Angsuran</b>', s_sm), Paragraph('<b>Tgl. Bayar</b>', s_sm), Paragraph('<b>Nilai (Rp.)</b>', s_sm)],
        ['UTJ / Booking Fee', f"{tgl_utj.day} {bulan_indo[tgl_utj.month]} {tgl_utj.year}", fmt(utj)],
    ]
    
    if len(cicilan_list) <= 6:
        for i, c in enumerate(cicilan_list):
            tgl = c.tanggal_jatuh_tempo
            schedule_data.append([f"Angsuran ke-{i+1}", f"{tgl.day} {bulan_indo[tgl.month]} {tgl.year}", fmt(c.jumlah_cicilan)])
    else:
        for i in range(5):
            c = cicilan_list[i]
            tgl = c.tanggal_jatuh_tempo
            schedule_data.append([f"Angsuran ke-{i+1}", f"{tgl.day} {bulan_indo[tgl.month]} {tgl.year}", fmt(c.jumlah_cicilan)])
        last = cicilan_list[-1]
        tgl_last = last.tanggal_jatuh_tempo
        schedule_data.append([f"... s/d Angsuran ke-{len(cicilan_list)}", f"{tgl_last.day} {bulan_indo[tgl_last.month]} {tgl_last.year}", fmt(last.jumlah_cicilan)])
    
    t_left = Table(schedule_data, colWidths=[3.2*cm, 2.5*cm, 2.5*cm])
    t_left.setStyle(TableStyle([
        ('FONT', (0,0), (-1,-1), 'Helvetica', 7),
        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.Color(0.92, 0.92, 0.92)),
        ('ALIGN', (2,0), (2,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    
    # RIGHT TABLE: Ringkasan
    summary_data = [
        [Paragraph('<b>Angsuran</b>', s_sm), Paragraph('<b>Tgl. Bayar</b>', s_sm), Paragraph('<b>Nilai (Rp.)</b>', s_sm)],
        ['Total DP', '', fmt(total_dp)],
    ]
    
    if customer.metode_pembelian == 'KPR':
        last_c = cicilan_list[-1] if cicilan_list else None
        tgl_pel = last_c.tanggal_jatuh_tempo if last_c else today
        summary_data.append([
            'Pelunasan KPR',
            f"{tgl_pel.day} {bulan_indo[tgl_pel.month]} {tgl_pel.year}",
            fmt(total_sisa_cicilan)
        ])
    
    summary_data.append([Paragraph('<b>Total</b>', s_sm), '', Paragraph(f'<b>{fmt(harga_jual_total)}</b>', s_sm)])
    
    t_right = Table(summary_data, colWidths=[2.5*cm, 2.5*cm, 2.8*cm])
    t_right.setStyle(TableStyle([
        ('FONT', (0,0), (-1,-1), 'Helvetica', 7),
        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.Color(0.92, 0.92, 0.92)),
        ('BACKGROUND', (0,-1), (-1,-1), colors.Color(0.95, 0.95, 0.95)),
        ('ALIGN', (2,0), (2,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    
    combined = Table([[t_left, t_right]], colWidths=[8.5*cm, 8*cm])
    combined.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    el.append(combined)
    el.append(Spacer(1, 3))
    
    if company.catatan_surat:
        el.append(Paragraph(f"<b>Catatan:</b> {company.catatan_surat}", s_n))
        el.append(Spacer(1, 4))
    
    # === C. PEMBAYARAN ===
    el.append(Paragraph("<b>C. PEMBAYARAN</b>", s_sec))
    el.append(Paragraph("Semua pembayaran melalui TRANSFER BANK dengan keterangan sebagai berikut:", s_n))
    el.append(Spacer(1, 2))
    
    el.append(Paragraph(f"<b>{company.nama_perusahaan}</b>", s_center))
    if company.rekening_1_bank and company.rekening_1_nomor:
        rek1 = f"A/C: {company.rekening_1_nomor} {company.rekening_1_bank}"
        if company.rekening_1_cabang:
            rek1 += f" {company.rekening_1_cabang}"
        el.append(Paragraph(rek1, s_center))
    if company.rekening_2_bank and company.rekening_2_nomor:
        rek2 = f"A/C: {company.rekening_2_nomor} {company.rekening_2_bank}"
        if company.rekening_2_cabang:
            rek2 += f" {company.rekening_2_cabang}"
        el.append(Paragraph(rek2, s_center))
    
    el.append(Spacer(1, 25))
    
    # === SIGNATURE (lebih longgar) ===
    sig_data = [
        [Paragraph('<b>Pembeli</b>', s_n), '', Paragraph('<b>Penjual</b>', s_n)],
        ['', '', ''],
        ['', '', ''],
        ['', '', ''],
        ['', '', ''],
        [Paragraph(f'<b>{customer.nama_lengkap}</b>', s_n), '', Paragraph(f'<b>{company.nama_perusahaan}</b>', s_n)],
    ]
    t_sig = Table(sig_data, colWidths=[6*cm, 4*cm, 6*cm])
    t_sig.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
    ]))
    el.append(t_sig)
    
    # Build PDF
    doc.build(el)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    nama_file = f"Surat_Pesanan_{customer.nama_lengkap.replace(' ', '_')}_{unit.kode_blok}.pdf"
    response['Content-Disposition'] = f'inline; filename="{nama_file}"'
    return response

